import os
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

def read_image_names(excel_path):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read the image names from the first column
    image_names = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0]:
            image_names.append(row[0])
    return image_names

def insert_images_to_excel(image_folder, image_names, output_excel_path):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set the height and width for the cells
    cell_size = 60
    for row in range(1, len(image_names) + 1):
        ws.row_dimensions[row].height = cell_size
        ws.column_dimensions['A'].width = cell_size / 5  # Adjusted for pixel to width ratio

    # Insert images and names
    for i, image_name in enumerate(image_names):
        image_path = os.path.join(image_folder, image_name)
        if os.path.exists(image_path):
            row = i + 1
            # Resize the image to fit into the cell
            img = PILImage.open(image_path)
            img.thumbnail((cell_size, cell_size), PILImage.LANCZOS)
            img.save(image_path)
            
            # Insert the image into the worksheet
            img = Image(image_path)
            img.anchor = f'A{row}'
            ws.add_image(img)
            
            # Insert the image name in the adjacent cell
            ws.cell(row=row, column=2, value=image_name)
        else:
            print(f"Image not found: {image_name}")

    # Save the workbook
    wb.save(output_excel_path)

# Example usage
image_folder = r'C:\Users\Qbits\Downloads\Rona 300 x 300 16MB\Rona 300 x 300 16MB - Copy'
input_excel_path = r'D:\COL\Target Image\Autamation Scripts\Image_Names.xlsx'
output_excel_path = 'output.xlsx'

# Read image names from the input Excel file
image_names = read_image_names(input_excel_path)

# Insert images into the new Excel file
insert_images_to_excel(image_folder, image_names, output_excel_path)