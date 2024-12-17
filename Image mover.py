import os
import shutil
import openpyxl

# Path to the Excel file containing the list of image names
excel_file = r"D:\COL\COL SUBMISSIONs\Versailles 10-30-24 New Products\Book1.xlsx"

# Path to the folder containing all images
all_images_folder = r"D:\COL\COL SUBMISSIONs\Versailles 10-30-24 New Products\Images"

# Read image names from Excel
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Find the column index of "image_names"
image_names_column_index = None
for col in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=col).value == "image_names":
        image_names_column_index = col
        break

if image_names_column_index is None:
    print("Column 'image_names' not found in the Excel file.")
    exit()

# Extract image names from the column
image_names = [sheet.cell(row=row, column=image_names_column_index).value for row in range(2, sheet.max_row + 1)]

# Iterate over each image name
for name in image_names:
    # Create folder for the current image name if it doesn't exist
    folder_path = os.path.join(os.getcwd(), name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    # Search for images containing the current name in the "All Images" folder
    for root, dirs, files in os.walk(all_images_folder):
        for file in files:
            if name in file:
                # Move or copy the matching image to the folder
                image_path = os.path.join(root, file)
                shutil.copy(image_path, folder_path)  # Change to shutil.move if you want to move instead of copy
